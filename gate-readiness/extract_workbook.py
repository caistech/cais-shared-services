#!/usr/bin/env python3
"""
Extract the ratified GATE_READINESS_REVIEW workbook into canonical JSON.

Single source of truth for the two fold-in destinations:
  - THIN_MVP_RUBRIC.md v2 (the doc)
  - the cockpit `readiness_criteria` + `promise_attributes` tables (the seed)

Run from this directory with the signed-off workbook in ~/Downloads/:
  python extract_workbook.py [path-to-xlsx]

Emits, next to this script:
  - criteria.json            (List 4 — 45 readiness checks, ratified tiers + weights)
  - promise-attributes.json  (List 3 — per-product "X, not Y" quality bars)
  - readiness_seed.sql       (idempotent UPSERT seed for both cockpit tables)
"""
import json
import os
import sys
import re

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(
    os.path.expanduser("~"), "Downloads", "GATE_READINESS_REVIEW_V4.xlsx"
)

# Workbook product-tab name -> cockpit methodology_hypothesis_cards.product_slug
# (matches constants.ts PLATFORMS[].slug / the seeded card slugs).
SLUG_MAP = {
    "Singify": "singify",
    "Connexions": "connexions",
    "Kira": "kira",
    "RaiseReady": "raiseready-template",
    "LingoPure": "lingopure-ai",
    "DealFindrs": "dealfindrs",
    "InvestorPilot": "investor-pilot",
    "PartnerPilot": "partner-pilot",
    "NDIS SDA": "ndis-sda-automate",
    "R&D Tax": "rnd-tax-tracker",
    "CQR": "cqr",
    "rehearsals-ai": "rehearsals-ai",
    "LaunchReady": "launchready",
    "UniversalLingo": "universallingo",
    "TourLingo": "tourlingo",
    "OutreachReady": "outreachready",
    "TenderWatch": "tenderwatch",
    "F2K Checkpoint": "checkpoint",
    "Disaster Support": "disaster-support",
    "Easy Claude Code": "easy-claude-code",
    # template-only tabs (no quality bars yet) — mapped for completeness
    "Universal Interviews": "universal-interviews",
    "mova": "mova",
    "storyverse": "storyverse",
    "F2K Projects": "f2k-projects",
    "LGA Planning DB": "lga-planning-db",
    "LeadSpark": "leadspark",
    "F2K Fund Token": "f2k-fund-tokenisation",
    "AIFTIS": "aiftis",
    "Lessons Learned": "lessonslearned",
    "Storefront MCP": "storefront-mcp",
    "Preflight": "preflight",
    "SmartBoard": "smartboard",
    "Hair Stylist AI": "hairstylist-ai",
    "OMQ Offshore": "f2k-offshore-modular",
}

NON_PRODUCT = {"README", "Criteria Classification"}


def s(v):
    return "" if v is None else str(v).strip()


def normalise_dashes(text):
    # the workbook uses a hyphen for the "X - not Y" split; keep as-is but
    # collapse the mojibake the source picked up from copy/paste.
    return text.replace("’", "'").replace("�", "-")


def extract_criteria(wb):
    ws = wb["Criteria Classification"]
    rows = list(ws.iter_rows(values_only=True))
    # header row 0: # | Check | Source | Method | Relevance | PROPOSED TIER |
    #   Confirm? | Your Tier | Weight | Notes
    out = []
    for r in rows[1:]:
        code = s(r[0])
        if not code:
            continue
        proposed = s(r[5])
        override = s(r[7])
        final_tier = override or proposed
        out.append(
            {
                "code": code,                       # P1..P4, 1..41
                "check": normalise_dashes(s(r[1])),
                "source": s(r[2]),
                "method": s(r[3]),
                "relevance": s(r[4]),
                "proposed_tier": proposed,
                "confirmed": s(r[6]).upper().startswith("Y"),
                "tier": final_tier,                 # override wins; here == proposed
                "weight": s(r[8]) or None,
                "notes": normalise_dashes(s(r[9])),
            }
        )
    return out


def extract_promises(wb):
    products = []
    for name in wb.sheetnames:
        if name in NON_PRODUCT:
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))

        def g(ri, ci):
            try:
                return s(rows[ri][ci])
            except IndexError:
                return ""

        needed = g(0, 1)
        promise = g(1, 1)
        distributor = g(2, 1)
        attrs = []
        for ri in range(4, len(rows)):
            attr = g(ri, 0)
            if not attr:
                continue
            claude_draft = normalise_dashes(g(ri, 1))
            your_bar = normalise_dashes(g(ri, 4))
            attrs.append(
                {
                    "attribute": attr,
                    "quality_bar": your_bar or None,         # ratified bar (None = not set)
                    "claude_draft": claude_draft or None,
                    "verify": g(ri, 2),
                    "approved": g(ri, 3).upper().startswith("Y"),
                    "notes": normalise_dashes(g(ri, 5)),
                }
            )
        filled = sum(1 for a in attrs if a["quality_bar"])
        products.append(
            {
                "tab": name,
                "slug": SLUG_MAP.get(name),
                "needed": needed or None,
                "promise": normalise_dashes(promise),
                "distributor": normalise_dashes(distributor) or None,
                "attributes": attrs,
                "bars_filled": filled,
            }
        )
    return products


def sql_lit(v):
    if v is None:
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def emit_seed_sql(criteria, products):
    lines = [
        "-- AUTO-GENERATED by gate-readiness/extract_workbook.py from the",
        "-- signed-off GATE_READINESS_REVIEW_V4 workbook. Do not hand-edit;",
        "-- re-run the extractor and copy the output if the workbook changes.",
        "",
        "-- readiness_criteria: the ratified 45-check catalogue (List 4).",
        "INSERT INTO readiness_criteria",
        "  (code, check_label, source, method, relevance, tier, weight, notes, sort_order)",
        "VALUES",
    ]
    vals = []
    for i, c in enumerate(criteria):
        vals.append(
            "  (%s, %s, %s, %s, %s, %s, %s, %s, %d)"
            % (
                sql_lit(c["code"]),
                sql_lit(c["check"]),
                sql_lit(c["source"]),
                sql_lit(c["method"]),
                sql_lit(c["relevance"]),
                sql_lit(c["tier"]),
                sql_lit(c["weight"]),
                sql_lit(c["notes"]),
                i,
            )
        )
    lines.append(",\n".join(vals))
    lines += [
        "ON CONFLICT (code) DO UPDATE SET",
        "  check_label = EXCLUDED.check_label, source = EXCLUDED.source,",
        "  method = EXCLUDED.method, relevance = EXCLUDED.relevance,",
        "  tier = EXCLUDED.tier, weight = EXCLUDED.weight,",
        "  notes = EXCLUDED.notes, sort_order = EXCLUDED.sort_order;",
        "",
        "-- promise_attributes: the per-product 'X, not Y' quality bars (List 3),",
        "-- only the products whose bars are ratified (bars_filled > 0).",
        "INSERT INTO promise_attributes",
        "  (product_slug, promise, distributor, attribute, quality_bar, verify, sort_order)",
        "VALUES",
    ]
    pvals = []
    for p in products:
        if not p["bars_filled"] or not p["slug"]:
            continue
        order = 0
        for a in p["attributes"]:
            if not a["quality_bar"]:
                continue
            pvals.append(
                "  (%s, %s, %s, %s, %s, %s, %d)"
                % (
                    sql_lit(p["slug"]),
                    sql_lit(p["promise"]),
                    sql_lit(p["distributor"]),
                    sql_lit(a["attribute"]),
                    sql_lit(a["quality_bar"]),
                    sql_lit(a["verify"]),
                    order,
                )
            )
            order += 1
    lines.append(",\n".join(pvals))
    lines += [
        "ON CONFLICT (product_slug, attribute) DO UPDATE SET",
        "  promise = EXCLUDED.promise, distributor = EXCLUDED.distributor,",
        "  quality_bar = EXCLUDED.quality_bar, verify = EXCLUDED.verify,",
        "  sort_order = EXCLUDED.sort_order;",
        "",
    ]
    return "\n".join(lines)


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    criteria = extract_criteria(wb)
    products = extract_promises(wb)

    tier_tally = {}
    for c in criteria:
        tier_tally[c["tier"]] = tier_tally.get(c["tier"], 0) + 1

    crit_doc = {
        "source": "GATE_READINESS_REVIEW_V4.xlsx — Criteria Classification (List 4)",
        "ratified": True,
        "all_confirmed": all(c["confirmed"] for c in criteria),
        "count": len(criteria),
        "tier_tally": tier_tally,
        "criteria": criteria,
    }
    prom_doc = {
        "source": "GATE_READINESS_REVIEW_V4.xlsx — Promise Attributes (List 3)",
        "products_with_bars": sum(1 for p in products if p["bars_filled"]),
        "products_total": len(products),
        "products": products,
    }

    with open(os.path.join(HERE, "criteria.json"), "w", encoding="utf-8") as f:
        json.dump(crit_doc, f, indent=2, ensure_ascii=False)
    with open(os.path.join(HERE, "promise-attributes.json"), "w", encoding="utf-8") as f:
        json.dump(prom_doc, f, indent=2, ensure_ascii=False)
    with open(os.path.join(HERE, "readiness_seed.sql"), "w", encoding="utf-8") as f:
        f.write(emit_seed_sql(criteria, products))

    print("criteria:", len(criteria), "confirmed:", crit_doc["all_confirmed"])
    print("tier tally:", tier_tally)
    print("products with bars:", prom_doc["products_with_bars"], "of", len(products))


if __name__ == "__main__":
    main()
